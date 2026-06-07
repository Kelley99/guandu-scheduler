#!/usr/bin/env python3
"""
官渡自动分配工具 - Flask后端
"""

import os
import re
import json
import random
import sqlite3
from datetime import datetime
from flask import send_file, Flask, render_template, request, jsonify, send_file, session, Response, make_response
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# OCR 懒加载
_easyocr_reader = None

# 腾讯云 OCR
import base64
from tencentcloud.common import credential
from tencentcloud.ocr.v20181119 import ocr_client, models

TENCENT_SECRET_ID = 'AKID_REMOVED_FROM_HISTORY'
TENCENT_SECRET_KEY = 'SECRET_KEY_REMOVED_FROM_HISTORY'

def tencent_ocr(image_path):
    """
    调用腾讯云通用文字识别API
    返回: 类似 EasyOCR 的格式 [(bbox, text, confidence), ...]
    """
    try:
        cred = credential.Credential(TENCENT_SECRET_ID, TENCENT_SECRET_KEY)
        client = ocr_client.OcrClient(cred, 'ap-guangzhou')

        with open(image_path, 'rb') as f:
            img_b64 = base64.b64encode(f.read()).decode('utf-8')

        req = models.GeneralBasicOCRRequest()
        req.ImageBase64 = img_b64

        resp = client.GeneralBasicOCR(req)

        results = []
        for item in resp.TextDetections:
            poly = item.Polygon
            if poly:
                x_coords = [p.X for p in poly]
                y_coords = [p.Y for p in poly]
                bbox = [[min(x_coords), min(y_coords)],
                        [max(x_coords), min(y_coords)],
                        [max(x_coords), max(y_coords)],
                        [min(x_coords), max(y_coords)]]
            else:
                bbox = [[0,0],[0,0],[0,0],[0,0]]
                print(f'[OCR Debug] 文字: {item.DetectedText}, 置信度: {item.Confidence}', flush=True)
            results.append((bbox, item.DetectedText, item.Confidence))
        print(f'[OCR Debug] 共识别 {len(results)} 个文字块', flush=True)
        return results
    except Exception as e:
        print(f'[OCR] 腾讯云OCR失败: {e}，尝试百度OCR兜底...')
        return baidu_ocr(image_path)


def baidu_ocr(image_path):
    """百度OCR通用文字识别（腾讯云失败时的兜底方案）"""
    try:
        import base64, requests, json, time
        
        API_KEY = 'OS2wp5hlvvJwJIYg5ayRA8kt'
        SECRET_KEY = 'VkbZhazXFLM3hswEtikSIiKGUOEpG1Ts'
        
        # 1. 获取 access_token
        token_url = f'https://aip.baidubce.com/oauth/2.0/token?grant_type=client_credentials&client_id={API_KEY}&client_secret={SECRET_KEY}'
        token_resp = requests.get(token_url, timeout=10)
        access_token = token_resp.json().get('access_token')
        if not access_token:
            print('[OCR] 百度OCR获取token失败')
            return []
        
        # 2. 调用通用文字识别接口
        ocr_url = f'https://aip.baidubce.com/rest/2.0/ocr/v1/general_basic?access_token={access_token}'
        with open(image_path, 'rb') as f:
            img_b64 = base64.b64encode(f.read()).decode('utf-8')
        data = {'image': img_b64}
        resp = requests.post(ocr_url, data=data, timeout=30)
        result = resp.json()
        
        if 'words_result' not in result:
            print(f'[OCR] 百度OCR返回错误: {result}')
            return []
        
        # 转换为类似 EasyOCR 的格式: [(bbox, text, confidence), ...]
        results = []
        for item in result['words_result']:
            words = item['words']
            # 百度OCR不返回bbox，用占位符
            bbox = [[0,0],[0,0],[0,0],[0,0]]
            confidence = 99  # 百度不返回confidence，默认99
            results.append((bbox, words, confidence))
        return results
    except Exception as e:
        print(f'[OCR] 百度OCR失败: {e}')
        return []

def get_ocr_reader():
    global _easyocr_reader
    if _easyocr_reader is None:
        import easyocr
        print('[OCR] 初始化EasyOCR reader (首次较慢)...')
        _easyocr_reader = easyocr.Reader(['ch_sim', 'en'], verbose=False)
        print('[OCR] Reader初始化完成')
    return _easyocr_reader

app = Flask(__name__)
app.secret_key = 'guandu-secret-key-2024'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

# CORS 头
@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    return response


# 数据目录
DATA_DIR = os.path.join(os.path.dirname(__file__), 'knowledge-base')
STATS_FILE = os.path.join(DATA_DIR, '凌霄数据统计表26.3.30.md')
GUANDU_FILE = os.path.join(DATA_DIR, '凌霄官渡26.md')

# 上传目录
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)

# 考勤数据库
ATTENDANCE_DB = os.path.join(os.path.dirname(__file__), 'gc_attendance.db')

def init_attendance_db():
    """初始化考勤数据库

    表结构:
    - members: 成员名单(从属性表导入)
    - assignments: 每周分配记录(关联成员与军团)
    - attendance_reports: 战报记录(每次战报一条)
    - attendance_detail: 考勤明细(成员+战报+状态)
    """
    conn = sqlite3.connect(ATTENDANCE_DB)
    c = conn.cursor()

    # 成员表:存储成员基础数据,可从属性表导入
    c.execute('''
        CREATE TABLE IF NOT EXISTS members (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 分配记录表:记录每周分配结果(关联哪个军团、哪个section)
    c.execute('''
        CREATE TABLE IF NOT EXISTS assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            section TEXT NOT NULL,          -- "团一" / "团二" / "团二 (26.4.5)" 等
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(section)
        )
    ''')

    # 分配明细表:每个分配记录包含哪些成员(正式/候补)
    c.execute('''
        CREATE TABLE IF NOT EXISTS assignment_members (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            assignment_id INTEGER NOT NULL,
            member_name TEXT NOT NULL,
            role TEXT DEFAULT '正式',         -- "正式" / "候补"
            position TEXT,                   -- 如 "B1", "D1-8", "候补"
            FOREIGN KEY (assignment_id) REFERENCES assignments(id),
            UNIQUE(assignment_id, member_name)
        )
    ''')

    # 战报记录表:每次战报一条
    c.execute('''
        CREATE TABLE IF NOT EXISTS attendance_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            section TEXT NOT NULL,          -- "团一" / "团二"
            report_date TEXT NOT NULL,      -- 如 "2026-05-10"
            screenshot TEXT,                -- 截图文件名
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 考勤明细表:每个成员在每次战报中的状态
    c.execute('''
        CREATE TABLE IF NOT EXISTS attendance_detail (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_id INTEGER NOT NULL,
            member_name TEXT NOT NULL,
            status TEXT DEFAULT '出席',      -- "出席" / "缺席" / "请假" / "候补上场"
            points INTEGER DEFAULT 0,       -- 功勋积分
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (report_id) REFERENCES attendance_reports(id),
            UNIQUE(report_id, member_name)
        )
    ''')

    conn.commit()
    conn.close()

def log_attendance(action: str, detail: str):
    """记录考勤操作日志(控制台)"""
    print(f'[考勤] {action}: {detail}')

# 启动时初始化数据库
init_attendance_db()
log_attendance('SYSTEM', '考勤数据库初始化完成')

# 允许的文件扩展名
ALLOWED_EXTENSIONS = {'xlsx', 'csv', 'md', 'txt'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ============ 解析函数 ============

def parse_stats_table(content: str) -> dict:
    """解析统计表,返回 {成员名: {hp, total, power, ...}}

    格式1: | 序号 | 成员名称 | 集结加成 | 步兵防御 | 步兵生命值 | ... | 六维属性总和 | ...
    格式2: | 序号 | 成员名称 | 战力 | ...
    支持动态定位 步兵生命值、六维属性总和、战力 列
    """
    stats = {}
    hp_col_idx = None
    total_col_idx = None
    power_col_idx = None

    for line in content.split('\n'):
        if '|' not in line:
            continue
        parts = [p.strip() for p in line.split('|')]
        parts = [p for p in parts if p]

        # 找表头确定各列索引
        if hp_col_idx is None and '步兵生命值' in parts:
            hp_col_idx = parts.index('步兵生命值')
        if total_col_idx is None and '六维属性总和' in parts:
            total_col_idx = parts.index('六维属性总和')
        if power_col_idx is None and '战力' in parts:
            power_col_idx = parts.index('战力')

        # 数据行
        if len(parts) >= 5 and parts[0].isdigit():
            name = parts[1]
            if name in stats:
                continue  # 避免重复

            hp = 0
            total = 0
            power = 0

            # 获取步兵生命值
            if hp_col_idx is not None and len(parts) > hp_col_idx:
                try:
                    hp = float(parts[hp_col_idx])
                except:
                    pass

            # 获取六维属性总和
            if total_col_idx is not None and len(parts) > total_col_idx:
                try:
                    total = float(parts[total_col_idx])
                except:
                    pass

            # 获取战力
            if power_col_idx is not None and len(parts) > power_col_idx:
                try:
                    power = float(parts[power_col_idx])
                except:
                    pass

            # 如果HP值异常(>100000),尝试智能检测
            if hp > 100000:
                for i, p in enumerate(parts):
                    try:
                        v = float(p)
                        if 100 <= v <= 5000:
                            hp = v
                            break
                    except:
                        pass

            # 如果还是异常,设为0
            if hp > 100000:
                hp = 0

            stats[name] = {
                'hp': hp,
                'total': total,
                'power': power,
            }

    return stats


def parse_guandu_table(content: str, section: str) -> dict:
    """解析官渡表,返回各队数据和候补名单

    表格结构:
    | 队伍 | 队长 | 队员 | 分组 | 队员1 | 战术1 | 战术2 |
    parts: ['', '队伍', '队长', '队员', '分组', '队员1', '战术1', '战术2']

    A/B行结构:
    | | | A | 队员名 |
    parts: ['', '', 'A', '队员名']
    """
    pattern = rf'## {re.escape(section)}[\s\S]*?(?=## |$)'
    match = re.search(pattern, content)
    if not match:
        return {'teams': {}, 'bench': []}

    section_content = match.group(0)
    lines = section_content.split('\n')

    tactic_kw = ['0-10分钟', '10-20分钟', '20分钟以后', '拿下', '驻守', '集结',
                 '粮仓', '乌巢', '官渡', '霹雳', '锱重', '驻防', '采集', '远程',
                 '不动', '应变', '必要时', '跟随', '工匠坊', '兵器坊', '首占']

    current_team = None
    teams_data = {}
    bench_members = []

    for line in lines:
        # 保留原始列位置,不过滤空字符串
        raw_parts = [p.strip() for p in line.split('|')]
        # 同时生成过滤版用于简单判断
        parts = [p for p in raw_parts if p]

        if not parts:
            continue

        # 跳过分隔符行
        if all(re.match(r'^-+$', p) for p in parts):
            continue

        # 队号行: parts[0] = '1队' 等
        # 原始列位置: col1=队伍, col2=队长, col3=分组(A/B/空), col4=队员, col5=0-10分钟, col6=10-20分钟, col7=20分钟以后
        if re.match(r'^\d+队$', parts[0]):
            current_team = parts[0]
            captain = raw_parts[2].strip() if len(raw_parts) > 2 else parts[1]
            teams_data[current_team] = {
                'captain': captain,
                'A_members': [],
                'B_members': [],
                'A_tasks': {'0-10': '', '10-20': '', '20+': ''},
                'B_tasks': {'0-10': '', '10-20': '', '20+': ''},
            }
            group = raw_parts[3].strip() if len(raw_parts) > 3 else ''
            member = raw_parts[4].strip() if len(raw_parts) > 4 else ''
            task_0_10 = raw_parts[5].strip() if len(raw_parts) > 5 else ''
            task_10_20 = raw_parts[6].strip() if len(raw_parts) > 6 else ''
            task_20_plus = raw_parts[7].strip() if len(raw_parts) > 7 else ''

            if group in ['A', 'B']:
                # 1队2队格式: 有AB分组
                if member and not any(kw in member for kw in tactic_kw):
                    teams_data[current_team][f'{group}_members'].append(member)
                teams_data[current_team][f'{group}_tasks']['0-10'] = task_0_10
                teams_data[current_team][f'{group}_tasks']['10-20'] = task_10_20
                teams_data[current_team][f'{group}_tasks']['20+'] = task_20_plus
            else:
                # 3-6队格式: 无AB分组,parts[2]=成员,parts[3]=时段1(10-20), parts[4]=时段2(20+)
                if member and not any(kw in member for kw in tactic_kw) and not member.isdigit():
                    teams_data[current_team]['A_members'].append(member)
                # 根据队类型取时段索引:1-2队用[4/5/6],3-6队用[3/4/5]
                if len(parts) > 3 and parts[2] in ['A', 'B']:
                    # 1队2队: [4]=0-10, [5]=10-20, [6]=20+
                    task_0_10 = parts[4].strip() if len(parts) > 4 else ''
                    task_10_20 = parts[5].strip() if len(parts) > 5 else ''
                    task_20_plus = parts[6].strip() if len(parts) > 6 else ''
                else:
                    # 3-6队: [3]=10-20(col5,含"大粮仓"), [4]=20+(col6), [5]=col7
                    task_0_10 = ''
                    task_10_20 = parts[3].strip() if len(parts) > 3 else ''
                    task_20_plus = parts[4].strip() if len(parts) > 4 else ''
                teams_data[current_team]['A_tasks']['0-10'] = task_0_10
                teams_data[current_team]['A_tasks']['10-20'] = task_10_20
                teams_data[current_team]['A_tasks']['20+'] = task_20_plus

        # A/B分组行: raw_parts[3] = 'A' 或 'B'
        elif len(raw_parts) > 4 and raw_parts[3].strip() in ['A', 'B'] and current_team:
            group = raw_parts[3].strip()
            member = raw_parts[4].strip() if len(raw_parts) > 4 else ''
            task_10_20 = raw_parts[6].strip() if len(raw_parts) > 6 else ''
            if member and not any(kw in member for kw in tactic_kw):
                teams_data[current_team][f'{group}_members'].append(member)
            # B组第一行可能带任务(如 col6='B队驻守')
            if group == 'B' and task_10_20 and any(kw in task_10_20 for kw in tactic_kw):
                teams_data[current_team]['B_tasks']['10-20'] = task_10_20

        # 无队号/分组标记的行,可能是队员行(3-6队后续队员)
        elif current_team and parts[0] not in ['队伍', '队长', '替补', '候补']:
            # 3-6队后续队员在 col4
            member = raw_parts[4].strip() if len(raw_parts) > 4 else parts[0]
            if member and not any(kw in member for kw in tactic_kw) and not member.isdigit():
                if not re.match(r'^[\d\s]+$', member):
                    teams_data[current_team]['A_members'].append(member)

        # 替补/候补行(支持表格格式 | 替补 | 名单 | 和非表格格式 候补：名单)
        bench_text = None
        if parts and parts[0] in ['替补', '候补'] and len(parts) > 1:
            # 表格格式: | 替补 | 名单 |
            bench_text = parts[1]
        elif line.startswith('候补') or line.startswith('替补'):
            # 非表格格式: 候补：名单 或 替补：名单（行首无 |）
            bench_text = line.replace('候补：', '').replace('替补：', '').strip()
        
        if bench_text is not None:
            bench_members = [m.strip() for m in bench_text.split('、') if m.strip()]
            # 替补任务在 raw_parts[6] (10-20分钟列)
            bench_task = raw_parts[6].strip() if len(raw_parts) > 6 else ''
            continue

    return {'teams': teams_data, 'bench': bench_members, 'bench_task': bench_task if 'bench_task' in dir() else ''}


def extract_j_members(teams_data: dict) -> list:
    """提取J列成员(队长+队员,去重)"""
    members = []
    seen = set()

    for team in ['1队', '2队', '3队', '4队', '5队', '6队']:
        if team in teams_data:
            data = teams_data[team]
            # 队长
            if data['captain'] and data['captain'] not in seen:
                seen.add(data['captain'])
                members.append(data['captain'])
            # A组成员
            for m in data['A_members']:
                for name in m.split('、'):
                    name = name.strip()
                    if name and name not in seen:
                        seen.add(name)
                        members.append(name)
            # B组成员
            for m in data['B_members']:
                for name in m.split('、'):
                    name = name.strip()
                    if name and name not in seen:
                        seen.add(name)
                        members.append(name)

    return members


def expand_members(members: list) -> list:
    """展开顿号分隔的成员名"""
    result = []
    seen = set()
    for m in members:
        if '、' in m:
            for name in m.split('、'):
                name = name.strip()
                if name and name not in seen:
                    seen.add(name)
                    result.append(name)
        else:
            if m and m not in seen:
                seen.add(m)
                result.append(m)
    return result


def parse_power_table(content: str) -> dict:
    """解析战力表,返回 {成员名: 战力值}

    支持格式:
    格式1: | 序号 | 成员名称 | 战力 | ...  (markdown表格)
    格式2: 纯文本 每行: 成员名 战力值
    """
    power_data = {}
    power_col_idx = None

    for line in content.split('\n'):
        if '|' not in line and not line.strip():
            continue

        # markdown表格格式
        if '|' in line:
            parts = [p.strip() for p in line.split('|')]
            parts = [p for p in parts if p]

            # 找表头
            if power_col_idx is None and '战力' in parts:
                power_col_idx = parts.index('战力')
                continue

            # 数据行
            if len(parts) >= 3 and parts[0].isdigit():
                name = parts[1]
                if power_col_idx is not None and len(parts) > power_col_idx:
                    try:
                        power = float(parts[power_col_idx])
                        if name and power > 0:
                            power_data[name] = power
                    except:
                        pass
                continue

        # 纯文本格式: 名字 数字
        stripped = line.strip()
        if stripped:
            # 匹配: 数字 开头(序号) 名字 数字(战力)
            m = re.match(r'^(\d+)\s+(\S+)\s+(\d[\d,]*\.?\d*)', stripped)
            if m:
                name = m.group(2)
                try:
                    power = float(m.group(3).replace(',', ''))
                    if name and power > 0:
                        power_data[name] = power
                except:
                    pass
                continue
            # 匹配: 名字 数字(战力)
            m = re.match(r'^(\S{2,10})\s+(\d[\d,]*\.?\d*)', stripped)
            if m:
                name = m.group(1)
                try:
                    power = float(m.group(2).replace(',', ''))
                    if name and power > 0:
                        power_data[name] = power
                except:
                    pass

    return power_data


@app.route('/api/sections')
def get_sections():
    """从默认官渡表提取可用分组列表"""
    try:
        with open(GUANDU_FILE, 'r', encoding='utf-8') as f:
            content = f.read()
        sections = []
        for m in re.finditer(r'^## (.+)$', content, re.MULTILINE):
            title = m.group(1).strip()
            if '排名' not in title:
                sections.append(title)
        return jsonify({'success': True, 'sections': sections})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/init')
def init_api():
    """一次性初始化:返回sections、stats、demo数据和匹配结果"""
    section = request.args.get('section', '团一')
    print(f'[INIT] section={repr(section)}')
    result = {'success': True}
    # 1. Sections
    try:
        with open(GUANDU_FILE, 'r', encoding='utf-8') as f:
            content = f.read()
        sections = []
        for m in re.finditer(r'^## (.+)$', content, re.MULTILINE):
            title = m.group(1).strip()
            if '排名' not in title:
                sections.append(title)
        result['sections'] = sections
    except Exception as e:
        app.logger.error(f'init sections error: {e}')
        result['sections'] = []

    # 2. Stats - 不再自动加载默认属性表,需用户手动上传或加载
    result['stats'] = {}
    result['statsCount'] = 0
    result['sortFields'] = []

    # 3. 人员名单 - 不再自动加载，用户需手动上传或输入
    result['members'] = []
    result['bench'] = []
    result['teams'] = {}

    # 4. Auto match - 不再自动匹配,需用户手动上传属性表后触发

    # Force Content-Length so axios doesn't hang on chunked transfer
    resp = make_response(json.dumps(result, ensure_ascii=False))
    resp.headers['Content-Type'] = 'application/json; charset=utf-8'
    resp.headers['Content-Length'] = len(resp.data)
    return resp


@app.route('/api/demo_data')
def get_demo_data():
    """从默认官渡表提取全部成员作为测试数据"""
    section = request.args.get('section', '团一')
    app.logger.info(f'demo_data: section={section}')
    try:
        with open(GUANDU_FILE, 'r', encoding='utf-8') as f:
            content = f.read()
        data = parse_guandu_table(content, section)
        members = extract_j_members(data['teams'])
        bench = data['bench']
        app.logger.info(f'demo_data: {len(members)} members, {len(bench)} bench')
        return jsonify({
            'success': True,
            'members': members,
            'bench': bench,
            'teams': data['teams'],
            'section': section
        })
    except Exception as e:
        app.logger.error(f'demo_data error: {e}')
        return jsonify({'success': False, 'error': str(e)})


def match_member(name: str, stats: dict) -> dict:
    """匹配成员名到统计表"""
    # 直接匹配
    if name in stats:
        return {'original': name, 'matched': name, 'hp': stats[name]['hp'], 'status': 'exact'}

    # 顿号/逗号转换匹配
    for stat_name in stats.keys():
        if name.replace('、', '丶') == stat_name or name.replace('、', ',') == stat_name:
            return {'original': name, 'matched': stat_name, 'hp': stats[stat_name]['hp'], 'status': 'exact'}

    # 部分匹配
    for stat_name in stats.keys():
        if name in stat_name or stat_name in name:
            if name.replace('、', '丶') in stat_name or stat_name.replace('丶', '、') in name:
                return {'original': name, 'matched': stat_name, 'hp': stats[stat_name]['hp'], 'status': 'partial'}

    # 未找到
    return {'original': name, 'matched': None, 'hp': 0, 'status': 'not_found'}


def assign_members(members: list, stats: dict, name_map: dict, threshold: float = 900, seed: int = 42, manual_captains: dict = None, sort_by: str = 'hp') -> dict:
    """按Kelley规则分配成员到B列和D列

    分配规则:
    - B1=第1名, B2=第2名
    - D1=第3名(1队A组), D9=第4名(1队B组)
    - B3-B6=第5-8名
    - 第9名开始蛇形分配到1-6队D位(每队4个D位循环)
    - <threshold的跳过1-2队,直接去3-6队
    """
    random.seed(seed)
    manual_captains = manual_captains or {}
    sort_key = sort_by if sort_by in ('hp', 'total', 'power') else 'hp'

    # 匹配成员并收集所有属性
    matched_members = []
    for m in members:
        mapped_name = name_map.get(m, m)
        if mapped_name and mapped_name in stats:
            hp = stats[mapped_name].get('hp', 0)
            total = stats[mapped_name].get('total', 0)
            power = stats[mapped_name].get('power', 0)
        else:
            hp = total = power = 0
        matched_members.append({
            'original': m, 'mapped': mapped_name if mapped_name else m,
            'hp': hp, 'total': total, 'power': power,
        })

    sorted_members = sorted(matched_members, key=lambda x: x[sort_key], reverse=True)
    b_assign = {}
    d_assign = {}

    # 1. 手动指定队长
    manual_used = set()
    for pos in ['B1', 'B2', 'B3', 'B4', 'B5', 'B6']:
        if manual_captains.get(pos):
            manual_name = manual_captains[pos]
            for m in members:
                if m == manual_name:
                    mapped = name_map.get(m, m)
                    s = stats.get(mapped, {})
                    found = {'original': m, 'mapped': mapped or m,
                             'hp': s.get('hp', 0), 'total': s.get('total', 0), 'power': s.get('power', 0)}
                    b_assign[pos] = found
                    manual_used.add(manual_name)
                    break

    auto_pool = [m for m in sorted_members if m['original'] not in manual_used]

    # D列位置: 1队A(D1-D4), 2队A(D5-D8), 1队B(D9-D12), 2队B(D13-D16), 3队(D17-D18), 4队(D19-D20), 5队(D21-D22), 6队(D23-D24)
    # 1队: D1-D4(位置0-3) + D9-D12(位置8-11)
    # 2队: D5-D8(位置4-7) + D13-D16(位置12-15)
    # 3队: D17-D18(位置16-17) / 4队: D19-D20(位置18-19) / 5队: D21-D22(位置20-21) / 6队: D23-D24(位置22-23)
    all_d = [
        'D1','D2','D3','D4',    # 1队A组 位置0-3
        'D5','D6','D7','D8',         # 2队A组 位置4-7
        'D9','D10','D11','D12',      # 1队B组 位置8-11
        'D13','D14','D15','D16',     # 2队B组 位置12-15
        'D17','D18',                 # 3队 位置16-17
        'D19','D20',                 # 4队 位置18-19
        'D21','D22',                 # 5队 位置20-21
        'D23','D24',                 # 6队 位置22-23
    ]

    # 2. 自动分配队长: B1=第1名, B2=第2名
    pool_idx = 0
    for pos in ['B1', 'B2']:
        if pos not in b_assign:
            if pool_idx < len(auto_pool):
                b_assign[pos] = auto_pool[pool_idx]
                pool_idx += 1

    # 3. D1=第3名, D9=第4名 (1队A/B各第1人)
    if pool_idx < len(auto_pool):
        d_assign['D1'] = auto_pool[pool_idx]; pool_idx += 1
    if pool_idx < len(auto_pool):
        d_assign['D9'] = auto_pool[pool_idx]; pool_idx += 1

    # 4. B3-B6=第5-8名
    for pos in ['B3', 'B4', 'B5', 'B6']:
        if pos not in b_assign:
            if pool_idx < len(auto_pool):
                b_assign[pos] = auto_pool[pool_idx]
                pool_idx += 1

    # 剩余成员（第9名开始）
    remaining = auto_pool[pool_idx:]

    # 蛇形分配剩余22人 → 22个D位（D1/D9已预分配）
    # 蛇形顺序: 1队A→2队A→1队B→2队B→3队→4队→5队→6队→6队→5队→4队→3队
    snake_positions = [
        'D2','D5','D3','D6','D4','D7','D8',   # 1队A→2队A 蛇形（D1已占）
        'D10','D13','D11','D14','D12','D15','D16',  # 1队B→2队B 蛇形（D9已占）
        'D17','D19','D21','D23','D24','D22','D20','D18'  # 3-6队 蛇形
    ]
    for i, pos in enumerate(snake_positions):
        if i < len(remaining):
            d_assign[pos] = remaining[i]

    return {
        'b_assign': b_assign,
        'd_assign': d_assign,
        'sorted': sorted_members,
        'unmatched': [m for m in matched_members if m[sort_key] == 0],
    }


# ============ 路由 ============

@app.route('/formation')
def formation_page():
    """队形配置页面"""
    with open(os.path.join(os.path.dirname(__file__), 'templates', 'formation.html'), 'r', encoding='utf-8') as f:
        html = f.read()
    return html

@app.route('/')
def index():
    """主页"""
    with open(os.path.join(os.path.dirname(__file__), 'templates', 'index.html'), 'r', encoding='utf-8') as f:
        html = f.read()
    return html



def detect_sort_fields(stats: dict) -> list:
    """检测属性表中哪些排序字段有数据,返回可用字段列表

    检查逻辑:如果某字段在任一成员中有非零值,则认为该字段可用
    """
    fields = []
    has_hp = False
    has_total = False
    has_power = False
    for v in stats.values():
        if v.get('hp', 0) > 0:
            has_hp = True
        if v.get('total', 0) > 0:
            has_total = True
        if v.get('power', 0) > 0:
            has_power = True
    if has_hp:
        fields.append({'key': 'hp', 'label': '步兵生命值'})
    if has_total:
        fields.append({'key': 'total', 'label': '六维属性总和'})
    if has_power:
        fields.append({'key': 'power', 'label': '战力'})
    return fields


@app.route('/api/load_stats')
def load_stats():
    """加载统计表数据(从默认文件)"""
    try:
        with open(STATS_FILE, 'r', encoding='utf-8') as f:
            content = f.read()
        stats = parse_stats_table(content)
        # 动态检测可用的排序字段
        fields = detect_sort_fields(stats)
        return jsonify({'success': True, 'stats': stats, 'count': len(stats), 'sort_fields': fields})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/upload_stats', methods=['POST'])
def upload_stats():
    """上传属性表文件"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '没有上传文件'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': '没有选择文件'})

    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': '不支持的文件格式,请上传 .xlsx, .csv 或 .md 文件'})

    try:
        filename = secure_filename(file.filename)
        ext = filename.rsplit('.', 1)[1].lower()

        # 根据文件类型解析
        stats = {}

        if ext == 'md':
            content = file.read().decode('utf-8')
            stats = parse_stats_table(content)

        elif ext == 'csv':
            content = file.read().decode('utf-8')
            stats = parse_stats_csv(content)

        elif ext == 'xlsx':
            stats = parse_stats_xlsx(file)

        if not stats:
            return jsonify({'success': False, 'error': '未能解析到有效数据,请检查文件格式'})

        fields = detect_sort_fields(stats)
        return jsonify({
            'success': True,
            'stats': stats,
            'count': len(stats),
            'filename': filename,
            'sort_fields': fields
        })

    except Exception as e:
        return jsonify({'success': False, 'error': f'解析失败: {str(e)}'})


@app.route('/api/upload_power', methods=['POST'])
def upload_power():
    """上传战力表文件"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '没有上传文件'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': '没有选择文件'})

    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': '不支持的文件格式,请上传 .xlsx, .csv 或 .md 文件'})

    try:
        filename = secure_filename(file.filename)
        ext = filename.rsplit('.', 1)[1].lower()

        power_data = {}

        if ext == 'md':
            content = file.read().decode('utf-8')
            power_data = parse_power_table(content)

        elif ext == 'csv':
            content = file.read().decode('utf-8')
            # 复用 parse_power_table 处理 CSV
            power_data = parse_power_table(content)

        elif ext == 'xlsx':
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
            power_col_idx = None
            header = None
            for row in ws.iter_rows(values_only=True):
                if not header:
                    header = row
                    for i, col in enumerate(row):
                        if col and '战力' in str(col):
                            power_col_idx = i
                    continue
                if len(row) < 2:
                    continue
                name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                if not name or name.isdigit():
                    continue
                if power_col_idx is not None and len(row) > power_col_idx and row[power_col_idx]:
                    try:
                        power = float(row[power_col_idx])
                        if power > 0:
                            power_data[name] = power
                    except:
                        pass
            wb.close()

        if not power_data:
            return jsonify({'success': False, 'error': '未能解析到有效战力数据,请检查文件格式'})

        return jsonify({
            'success': True,
            'power_data': power_data,
            'count': len(power_data),
            'filename': filename
        })

    except Exception as e:
        return jsonify({'success': False, 'error': f'解析失败: {str(e)}'})


def parse_stats_csv(content: str) -> dict:
    """解析CSV格式的属性表"""
    import csv
    from io import StringIO

    stats = {}
    reader = csv.reader(StringIO(content))
    header = None
    hp_idx = None
    total_idx = None
    power_idx = None

    for row in reader:
        if not header:
            header = row
            for i, col in enumerate(row):
                if '步兵生命值' in col or '生命值' in col or 'HP' in col.upper():
                    hp_idx = i
                elif '六维属性总和' in col or '六维总和' in col:
                    total_idx = i
                elif '战力' in col:
                    power_idx = i
            continue

        if len(row) < 2:
            continue

        name = row[1].strip() if len(row) > 1 else row[0].strip()
        if not name or name.isdigit():
            continue

        hp = 0
        total = 0
        power = 0

        if hp_idx is not None and len(row) > hp_idx:
            try: hp = float(row[hp_idx])
            except: pass
        if total_idx is not None and len(row) > total_idx:
            try: total = float(row[total_idx])
            except: pass
        if power_idx is not None and len(row) > power_idx:
            try: power = float(row[power_idx])
            except: pass

        # 如果HP异常,尝试智能检测
        if hp > 100000 or (hp == 0 and total == 0 and power == 0):
            for val in row:
                try:
                    v = float(val)
                    if 100 <= v <= 5000:
                        hp = v
                        break
                except:
                    pass

        if hp > 100000:
            hp = 0

        stats[name] = {'hp': hp, 'total': total, 'power': power}

    return stats


def parse_stats_xlsx(file) -> dict:
    """解析Excel格式的属性表（支持属性表和成员表两种格式）"""
    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb.active

    stats = {}
    header = None
    hp_idx = None
    total_idx = None
    power_idx = None
    # 成员表格式：列B=姓名，列E=战力（无hp/total）
    is_member_table = False
    # 跳过标题行（如"表格 1"）
    row_count = 0

    for row in ws.iter_rows(values_only=True):
        row_count += 1
        if not header:
            header = row
            for i, col in enumerate(row):
                if col and ('步兵生命值' in str(col) or '生命值' in str(col) or 'HP' in str(col).upper()):
                    hp_idx = i
                elif col and ('六维属性总和' in str(col) or '六维总和' in str(col)):
                    total_idx = i
                elif col and '战力' in str(col):
                    power_idx = i
            # 判断是否是成员表格式（没有hp/total但有战力）
            if hp_idx is None and total_idx is None and power_idx is not None:
                is_member_table = True
            # 如果第一行没有找到关键列但第二行有，跳到第二行
            if power_idx is None and row_count == 1:
                header = None  # 继续读下一行
                continue
            continue

        if is_member_table:
            # 成员表格式：列B(索引1)=姓名，列E(索引4)=战力
            name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            if not name or name.isdigit():
                continue
            power = 0
            if power_idx is not None and len(row) > power_idx and row[power_idx]:
                try: power = float(row[power_idx])
                except: pass
            stats[name] = {'hp': 0, 'total': 0, 'power': power}
            continue

        if len(row) < 2:
            continue

        name = str(row[1]).strip() if len(row) > 1 and row[1] else str(row[0]).strip() if row[0] else ''
        if not name or name.isdigit():
            continue

        hp = 0
        total = 0
        power = 0

        if hp_idx is not None and len(row) > hp_idx and row[hp_idx]:
            try: hp = float(row[hp_idx])
            except: pass
        if total_idx is not None and len(row) > total_idx and row[total_idx]:
            try: total = float(row[total_idx])
            except: pass
        if power_idx is not None and len(row) > power_idx and row[power_idx]:
            try: power = float(row[power_idx])
            except: pass

        if hp > 100000 or (hp == 0 and total == 0 and power == 0):
            for val in row:
                try:
                    v = float(val) if val else 0
                    if 100 <= v <= 5000:
                        hp = v
                        break
                except:
                    pass

        if hp > 100000:
            hp = 0

        stats[name] = {'hp': hp, 'total': total, 'power': power}

    wb.close()
    return stats


@app.route('/api/load_guandu')
def load_guandu():
    """加载官渡表数据(从默认文件)"""
    section = request.args.get('section', '团一')
    try:
        with open(GUANDU_FILE, 'r', encoding='utf-8') as f:
            content = f.read()
        data = parse_guandu_table(content, section)
        members = extract_j_members(data['teams'])
        return jsonify({
            'success': True,
            'teams': data['teams'],
            'bench': data['bench'],
            'members': members,
            'bench_task': data.get('bench_task', ''),
            'section': section,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/upload_guandu', methods=['POST'])
def upload_guandu():
    """上传官渡名单文件"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '没有上传文件'})

    file = request.files['file']
    section = request.form.get('section', '团一')

    if file.filename == '':
        return jsonify({'success': False, 'error': '没有选择文件'})

    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': '不支持的文件格式,请上传 .xlsx, .csv 或 .md 文件'})

    try:
        filename = secure_filename(file.filename)
        ext = filename.rsplit('.', 1)[1].lower()

        teams_data = {}
        bench = []
        members = []

        if ext == 'md':
            content = file.read().decode('utf-8')
            data = parse_guandu_table(content, section)
            teams_data = data['teams']
            bench = data['bench']
            members = extract_j_members(teams_data)

        elif ext == 'csv':
            content = file.read().decode('utf-8')
            members, bench = parse_guandu_csv(content)

        elif ext == 'xlsx':
            members, bench, teams_data = parse_guandu_xlsx(file)
            app.logger.info(f'xlsx解析: members={len(members)}, bench={len(bench)}, teams={len(teams_data)}')

        elif ext == 'txt':
            content = file.read().decode('utf-8')
            members, bench = parse_guandu_txt(content)
            app.logger.info(f'txt解析: members={len(members)}, bench={len(bench)}')

        if not members and not teams_data:
            return jsonify({'success': False, 'error': '未能解析到有效数据,请检查文件格式'})

        return jsonify({
            'success': True,
            'teams': teams_data,
            'bench': bench,
            'members': members,
            'section': section,
            'filename': filename
        })

    except Exception as e:
        return jsonify({'success': False, 'error': f'解析失败: {str(e)}'})


def parse_guandu_csv(content: str) -> tuple:
    """解析CSV格式的官渡名单,返回 (members, bench)"""
    import csv
    from io import StringIO

    members = []
    bench = []

    reader = csv.reader(StringIO(content))
    for row in reader:
        for val in row:
            val = str(val).strip()
            if not val or val.isdigit():
                continue
            if '替补' in val:
                # 后面的值可能是替补名单
                continue
            if val in ['队伍', '队长', '队员', 'A', 'B', '分组']:
                continue
            # 展开顿号分隔的名字
            for name in val.replace(',', '、').split('、'):
                name = name.strip()
                if name and name not in members:
                    members.append(name)

    return members, bench


def parse_guandu_txt(content: str) -> tuple:
    """解析纯文本格式的官渡名单,返回 (members, bench)

    格式示例:
    正式成员:
    张三
    李四
    王五

    候补成员:
    赵六
    陈七
    """
    members = []
    bench = []

    current_section = None  # None | 'formal' | 'bench'
    for line in content.split('\n'):
        line = line.strip()
        if not line:
            continue

        # 检测section切换
        ll = line.lower()
        if '正式' in line or '正式成员' in line:
            current_section = 'formal'
            continue
        if '候补' in line or '替补' in line or '备选' in line:
            current_section = 'bench'
            continue

        # 第一行可能是标题行(只有两个字,如"正式成员:")
        if current_section is None:
            # 如果还没有识别到section,先尝试把第一行当作正式成员
            current_section = 'formal'

        # 过滤掉分隔线、序号等
        if line in ['─', '-', '--', '~~', '=='] or line.startswith('第') and ('队' in line or '组' in line):
            continue

        target = bench if current_section == 'bench' else members

        # 展开顿号/逗号分隔的名字
        for name in line.replace(',', '、').replace(',', '、').split('、'):
            name = name.strip()
            # 过滤空字符串和数字
            if name and not name.isdigit():
                target.append(name)

    return members, bench


def parse_guandu_xlsx(file) -> tuple:
    """解析Excel格式的官渡名单,返回 (members, bench, teams_data)"""
    wb = openpyxl.load_workbook(file, read_only=True)

    # 找到包含"队伍"表头的工作表
    target_ws = None
    for name in wb.sheetnames:
        ws_check = wb[name]
        first_row = [str(c).strip() for c in next(ws_check.iter_rows(max_row=1, values_only=True), [])]
        if '队伍' in first_row:
            target_ws = ws_check
            break
    if target_ws is None:
        target_ws = wb.active

    members = []
    bench = []
    teams_data = {}
    current_team = None
    seen_names = set()

    for row in target_ws.iter_rows(values_only=True):
        # 安全获取列
        col0 = str(row[0]).strip() if row[0] else ''  # A列: "1队", "2队", "替补"
        col1 = str(row[1]).strip() if row[1] else ''  # B列: 队长名
        col2 = str(row[2]).strip() if row[2] else ''  # C列: A/B分组
        col3 = str(row[3]).strip() if len(row) > 3 and row[3] else ''  # D列: 成员名

        # 任务列 (E-G, index 4-6)
        task_0_10 = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        task_10_20 = str(row[5]).strip() if len(row) > 5 and row[5] else ''
        task_20_plus = str(row[6]).strip() if len(row) > 6 and row[6] else ''

        # 识别替补行
        if col0 == '替补':
            for i in range(9, min(20, len(row))):
                val = row[i] if i < len(row) else None
                if val and str(val).strip():
                    v = str(val).strip()
                    if v not in ['替补进入后,哪里缺人去哪里', '替补']:
                        for name in v.replace(',', '、').split('、'):
                            name = name.strip()
                            if name and name not in bench:
                                bench.append(name)
            continue

        # 识别队伍行(新队伍开始)
        if col0 and col0.endswith('队'):
            current_team = col0
            teams_data[current_team] = {
                'A_members': [],
                'B_members': [],
                'A_tasks': {'0-10': '', '10-20': '', '20+': ''},
                'B_tasks': {'10-20': '', '20+': ''},
            }
            # 队长
            if col1 and col1 not in seen_names:
                members.append(col1)
                seen_names.add(col1)

            # 队长行的 col3 成员(第一个队员)
            if col3 and col3 not in seen_names:
                for name in col3.replace(',', '、').split('、'):
                    name = name.strip()
                    if name and name not in seen_names:
                        members.append(name)
                        seen_names.add(name)
                        if current_team in ['3队', '4队', '5队', '6队']:
                            teams_data[current_team]['A_members'].append(name)
                        elif col2 == 'A':
                            teams_data[current_team]['A_members'].append(name)
                        elif col2 == 'B':
                            teams_data[current_team]['B_members'].append(name)

            # 提取该队 A 组任务(从队长行的 E-G 列)
            if current_team in ['1队', '2队']:
                if task_0_10: teams_data[current_team]['A_tasks']['0-10'] = task_0_10
                if task_10_20: teams_data[current_team]['A_tasks']['10-20'] = task_10_20
                if task_20_plus: teams_data[current_team]['A_tasks']['20+'] = task_20_plus
            else:
                # 3-6队:只有 A 组,任务在 E 列(0-10分钟)
                if task_0_10: teams_data[current_team]['A_tasks']['0-10'] = task_0_10
                if task_10_20: teams_data[current_team]['A_tasks']['10-20'] = task_10_20
                if task_20_plus: teams_data[current_team]['A_tasks']['20+'] = task_20_plus
            continue

        # 跳过标题行
        if col0 == '队伍' or col0.startswith('团'):
            continue

        # 队员行 (D列有成员名)
        if col3:
            if col2 in ['A', 'B']:
                # 1-2队: 有A/B分组
                group = col2
                for name in col3.replace(',', '、').split('、'):
                    name = name.strip()
                    if name and name not in seen_names:
                        members.append(name)
                        seen_names.add(name)
                        if current_team:
                            if group == 'A':
                                teams_data[current_team]['A_members'].append(name)
                            elif group == 'B':
                                teams_data[current_team]['B_members'].append(name)
                
                # B 组任务(1-2队,从 B 组行的 F 列提取)
                if group == 'B' and current_team in ['1队', '2队']:
                    if task_10_20: teams_data[current_team]['B_tasks']['10-20'] = task_10_20
                    if task_20_plus: teams_data[current_team]['B_tasks']['20+'] = task_20_plus
            elif current_team in ['3队', '4队', '5队', '6队']:
                # 3-6队: 无A/B分组,直接加入A_members
                for name in col3.replace(',', '、').split('、'):
                    name = name.strip()
                    if name and name not in seen_names:
                        pass
                        members.append(name)
                        seen_names.add(name)
                        if current_team:
                            teams_data[current_team]['A_members'].append(name)

    wb.close()
    return members, bench, teams_data


@app.route('/api/match_members', methods=['POST'])
def match_members_api():
    """匹配成员"""
    app.logger.info('match_members 被调用')
    data = request.json
    members = data.get('members', [])
    stats = data.get('stats', {})
    app.logger.info(f'match_members: {len(members)} members, {len(stats)} stats')

    results = []
    for m in members:
        result = match_member(m, stats)
        results.append(result)

    unmatched = [r for r in results if r['status'] == 'not_found']
    matched = [r for r in results if r['status'] != 'not_found']

    return jsonify({
        'success': True,
        'results': results,
        'matched': matched,
        'unmatched': unmatched,
    })


@app.route('/api/assign', methods=['POST'])
def assign_api():
    """分配成员并保存分配记录"""
    data = request.json
    members = data.get('members', [])
    stats = data.get('stats', {})
    name_map = data.get('name_map', {})
    threshold = data.get('threshold', 900)
    seed = data.get('seed', 42)
    manual_captains = data.get('manual_captains', {})
    sort_by = data.get('sort_by', 'hp')  # 'hp', 'total', 或 'power'
    power_data = data.get('power_data', {})  # 独立战力表数据
    section = data.get('section', '团一')  # 所属军团
    bench_members = data.get('bench_members', [])  # 候补成员列表

    # 如果有独立战力表数据,合并到 stats 中
    if power_data:
        for name, power_val in power_data.items():
            if name in stats:
                stats[name]['power'] = power_val
            else:
                # 新成员(战力表有但属性表没有)
                stats[name] = {'hp': 0, 'total': 0, 'power': power_val}

    result = assign_members(members, stats, name_map, threshold, seed, manual_captains, sort_by)

    # 保存分配记录到数据库
    try:
        conn = sqlite3.connect(ATTENDANCE_DB)
        c = conn.cursor()
        c.execute('INSERT INTO assignments (section) VALUES (?)', (section,))
        conn.commit()
        assignment_id = c.lastrowid
        for m in result['sorted']:
            c.execute('INSERT OR IGNORE INTO assignment_members (assignment_id, member_name, role) VALUES (?,?,?)',
                      (assignment_id, m['original'], '正式'))
        conn.commit()
        conn.close()
        log_attendance('SAVE_ASSIGNMENT', f'section={section} id={assignment_id} formal={len(result["sorted"])} bench={len(bench_members)}')

        # 插入 assignment_members 表（候补成员）
        for bench_name in bench_members:
            c.execute('INSERT OR IGNORE INTO assignment_members (assignment_id, member_name, role) VALUES (?,?,?)',
                      (assignment_id, bench_name, '候补'))
    except Exception as e:
        print('保存分配记录失败:', e)

    return jsonify({
        'success': True,
        'b_assign': result['b_assign'],
        'd_assign': result['d_assign'],
        'sorted': result['sorted'],
        'unmatched': result['unmatched'],
    })


@app.route('/api/export', methods=['POST'])
def export_api():
    """导出Excel - 将分配结果填入官渡表格式"""
    data = request.json

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "官渡分配表"

    # 样式
    header_font = Font(bold=True, size=12)
    captain_font = Font(bold=True, size=11, color='1a1a2e')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    captain_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')

    b_assign = data.get('b_assign', {})
    d_assign = data.get('d_assign', {})
    section = data.get('section', '团一')
    sort_by = data.get('sort_by', 'hp')
    teams_data = data.get('teams_data', {})
    bench_task = data.get('bench_task', '')

    # 队伍到D位置的映射
    # 1队2队: A组4人 + B组4人 = 8行队员
    # 3-6队: 无分组, 2行队员
    team_d_map = {
        1: {'a': ['D1','D2','D3','D4'], 'b': ['D9','D10','D11','D12']},
        2: {'a': ['D5','D6','D7','D8'], 'b': ['D13','D14','D15','D16']},
        3: {'members': ['D17','D18']},
        4: {'members': ['D19','D20']},
        5: {'members': ['D21','D22']},
        6: {'members': ['D23','D24']},
    }

    # 写入表头(不含HP/排序值列)
    headers = ['队伍', 'B列(队长)', '分组', 'D列(队员)', '0-10分钟', '10-20分钟', '20分钟以后', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    def set_cell(ws, row, col, value='', font=None, fill=None, align=None, border=None):
        """写入单元格并设置样式"""
        cell = ws.cell(row=row, column=col, value=value)
        if font: cell.font = font
        if fill: cell.fill = fill
        if align: cell.alignment = align
        if border: cell.border = border
        return cell

    def apply_border_range(ws, r1, c1, r2, c2):
        """给范围内的所有单元格加边框"""
        for r in range(r1, r2+1):
            for c in range(c1, c2+1):
                ws.cell(row=r, column=c).border = thin_border

    row = 2
    # 3-5队跨队合并追踪
    group_35_start_row = None
    group_35_end_row = None
    group_35_gh_content = ''

    for team_num in range(1, 7):
        team_name = f'{team_num}队'
        b_key = f'B{team_num}'
        captain_data = b_assign.get(b_key, {})
        captain = captain_data.get('original', '') if captain_data else ''
        captain_score = captain_data.get(sort_by, 0) if captain_data else 0
        tmap = team_d_map.get(team_num, {})

        if team_num <= 2:
            # === 1队2队: 8行 - A组4队员 + B组4队员 ===
            a_keys = tmap.get('a', [])
            b_keys = tmap.get('b', [])
            tdata = teams_data.get(team_name, {})
            a_tasks = tdata.get('A_tasks', {})
            b_tasks = tdata.get('B_tasks', {})

            team_start_row = row
            a_start_row = row

            # --- 写入A组4行 ---
            for i, dk in enumerate(a_keys):
                d_data = d_assign.get(dk, {})
                m_name = d_data.get('original', '') if d_data else ''
                m_score = d_data.get(sort_by, 0) if d_data else 0
                set_cell(ws, row, 3, 'A', align=center_align, border=thin_border)
                set_cell(ws, row, 4, m_name, border=thin_border)
                set_cell(ws, row, 8, '', border=thin_border)
                row += 1
            a_end_row = row - 1

            b_start_row = row

            # --- 写入B组4行 ---
            for i, dk in enumerate(b_keys):
                d_data = d_assign.get(dk, {})
                m_name = d_data.get('original', '') if d_data else ''
                set_cell(ws, row, 3, 'B', align=center_align, border=thin_border)
                set_cell(ws, row, 4, m_name, border=thin_border)
                set_cell(ws, row, 8, '', border=thin_border)
                row += 1
            b_end_row = row - 1
            team_end_row = row - 1

            # --- 合并单元格:队伍名(全队8行) ---
            if team_end_row > team_start_row:
                ws.merge_cells(start_row=team_start_row, start_column=1, end_row=team_end_row, end_column=1)
            set_cell(ws, team_start_row, 1, team_name, font=Font(bold=True, size=12), align=center_align, border=thin_border)
            apply_border_range(ws, team_start_row, 1, team_end_row, 1)

            # --- 合并单元格:队长(全队8行) ---
            if team_end_row > team_start_row:
                ws.merge_cells(start_row=team_start_row, start_column=2, end_row=team_end_row, end_column=2)
            set_cell(ws, team_start_row, 2, captain, font=captain_font, fill=captain_fill, align=center_align, border=thin_border)
            apply_border_range(ws, team_start_row, 2, team_end_row, 2)

            # --- 合并单元格:0-10分钟任务(全队8行) ---
            task_010 = a_tasks.get('0-10', '')
            if team_end_row > team_start_row:
                ws.merge_cells(start_row=team_start_row, start_column=5, end_row=team_end_row, end_column=5)
            set_cell(ws, team_start_row, 5, task_010, align=Alignment(wrap_text=True, vertical='center'), border=thin_border)
            apply_border_range(ws, team_start_row, 5, team_end_row, 5)

            # --- 合并单元格:A组 10-20分钟(4行) ---
            if a_end_row > a_start_row:
                ws.merge_cells(start_row=a_start_row, start_column=6, end_row=a_end_row, end_column=6)
            set_cell(ws, a_start_row, 6, a_tasks.get('10-20', ''), align=Alignment(wrap_text=True, vertical='center'), border=thin_border)
            apply_border_range(ws, a_start_row, 6, a_end_row, 6)

            # --- 合并单元格:A组 20+分钟(4行) ---
            if a_end_row > a_start_row:
                ws.merge_cells(start_row=a_start_row, start_column=7, end_row=a_end_row, end_column=7)
            set_cell(ws, a_start_row, 7, a_tasks.get('20+', ''), align=Alignment(wrap_text=True, vertical='center'), border=thin_border)
            apply_border_range(ws, a_start_row, 7, a_end_row, 7)

            # --- 合并单元格:B组 10-20分钟(4行) ---
            if b_end_row > b_start_row:
                ws.merge_cells(start_row=b_start_row, start_column=6, end_row=b_end_row, end_column=6)
            set_cell(ws, b_start_row, 6, b_tasks.get('10-20', ''), align=Alignment(wrap_text=True, vertical='center'), border=thin_border)
            apply_border_range(ws, b_start_row, 6, b_end_row, 6)

            # --- 合并单元格:B组 20+分钟(4行) ---
            if b_end_row > b_start_row:
                ws.merge_cells(start_row=b_start_row, start_column=7, end_row=b_end_row, end_column=7)
            set_cell(ws, b_start_row, 7, b_tasks.get('20+', ''), align=Alignment(wrap_text=True, vertical='center'), border=thin_border)
            apply_border_range(ws, b_start_row, 7, b_end_row, 7)

        else:
            # === 3-6队: 队长+队员 ===
            member_keys = tmap.get('members', [])
            tdata = teams_data.get(team_name, {})
            a_tasks = tdata.get('A_tasks', {})

            # 收集有数据的队员
            filled_members = []
            for mk in member_keys:
                d_data = d_assign.get(mk, {})
                m_name = d_data.get('original', '') if d_data else ''
                if m_name:
                    filled_members.append((m_name, d_data.get(sort_by, 0) if d_data else 0))

            team_start_row = row
            # 写入队员行
            for i, (m_name, m_score) in enumerate(filled_members):
                set_cell(ws, row, 4, m_name, border=thin_border)
                set_cell(ws, row, 8, '', border=thin_border)
                row += 1
            team_end_row = row - 1

            # 如果没有队员,至少写一行
            if not filled_members:
                set_cell(ws, row, 4, '', border=thin_border)
                set_cell(ws, row, 8, '', border=thin_border)
                team_end_row = row
                row += 1

            # --- 合并单元格:队伍名 ---
            if team_end_row > team_start_row:
                ws.merge_cells(start_row=team_start_row, start_column=1, end_row=team_end_row, end_column=1)
            set_cell(ws, team_start_row, 1, team_name, font=Font(bold=True, size=12), align=center_align, border=thin_border)
            apply_border_range(ws, team_start_row, 1, team_end_row, 1)

            # --- 合并单元格:队长 ---
            if team_end_row > team_start_row:
                ws.merge_cells(start_row=team_start_row, start_column=2, end_row=team_end_row, end_column=2)
            set_cell(ws, team_start_row, 2, captain, font=captain_font, fill=captain_fill, align=center_align, border=thin_border)
            apply_border_range(ws, team_start_row, 2, team_end_row, 2)

            # --- 任务列 ---
            task_010 = a_tasks.get('0-10', '')
            task_1020 = a_tasks.get('10-20', '')
            task_20plus = a_tasks.get('20+', '')

            if team_num <= 5:
                # === 3-5队 ===
                # E列(0-10分钟):每队的任务相同,队内垂直合并
                f_content = task_1020  # 实际内容来自10-20分钟字段
                if f_content:
                    if team_end_row > team_start_row:
                        ws.merge_cells(start_row=team_start_row, start_column=5, end_row=team_end_row, end_column=5)
                    set_cell(ws, team_start_row, 5, f_content, align=Alignment(wrap_text=True, vertical='center'), border=thin_border)
                    apply_border_range(ws, team_start_row, 5, team_end_row, 5)

                # F:G列(10-20分钟+20分钟以后):3-5队任务一样,记录起止行,循环结束后统一跨队合并
                if team_num == 3:
                    group_35_start_row = team_start_row
                    group_35_gh_content = task_20plus  # 使用3队的20+任务内容
                if team_num == 5:
                    group_35_end_row = team_end_row

            else:
                # === 6队 ===
                # E:F合并(0-10+10-20分钟),内容=task_1020
                ws.merge_cells(start_row=team_start_row, start_column=5, end_row=team_end_row, end_column=6)
                set_cell(ws, team_start_row, 5, task_1020, align=Alignment(wrap_text=True, vertical='center'), border=thin_border)
                apply_border_range(ws, team_start_row, 5, team_end_row, 6)

                # G列(20分钟以后)合并,内容=task_20plus
                if team_end_row > team_start_row:
                    ws.merge_cells(start_row=team_start_row, start_column=7, end_row=team_end_row, end_column=7)
                set_cell(ws, team_start_row, 7, task_20plus, align=Alignment(wrap_text=True, vertical='center'), border=thin_border)
                apply_border_range(ws, team_start_row, 7, team_end_row, 7)

    # --- 3-5队 G:H跨队合并(10-20分钟+20分钟以后任务相同) ---
    if group_35_start_row and group_35_end_row:
        ws.merge_cells(start_row=group_35_start_row, start_column=6, end_row=group_35_end_row, end_column=7)
        set_cell(ws, group_35_start_row, 6, group_35_gh_content, align=Alignment(wrap_text=True, vertical='center'), border=thin_border)
        apply_border_range(ws, group_35_start_row, 6, group_35_end_row, 7)

    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 15

    # 候补人员
    bench_list = data.get('bench_list', [])
    if bench_list:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value='候补人员')
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill(start_color='6c757d', end_color='6c757d', fill_type='solid')
        cell.alignment = center_align
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thin_border
        row += 1
        # 候补人员名单合并到 B:D
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        bench_names = '、'.join(bench_list)
        cell_names = ws.cell(row=row, column=2, value=bench_names)
        cell_names.alignment = Alignment(wrap_text=True)
        cell_names.border = thin_border
        ws.cell(row=row, column=1).border = thin_border
        for col in range(5, 9):
            ws.cell(row=row, column=col).border = thin_border
        # 替补任务描述移到同一行 E:G
        if bench_task:
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
            cell_task = ws.cell(row=row, column=5, value=bench_task)
            cell_task.alignment = Alignment(wrap_text=True)
            cell_task.border = thin_border
        row += 1

    # 底部备注
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    note = ws.cell(row=row, column=1, value='*各自队伍拿下首占后,可根据对手的动向进行调整,随机应变。*')
    note.alignment = Alignment(horizontal='center', wrap_text=True)
    note.font = Font(italic=True, size=10, color='666666')

    # 保存
    filename = f'官渡分配表_{section}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(os.path.dirname(__file__), 'downloads', filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)

    return jsonify({'success': True, 'filename': filename, 'path': filepath})


@app.route('/download/<filename>')
def download(filename):
    """下载文件"""
    filepath = os.path.join(os.path.dirname(__file__), 'downloads', filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return 'File not found', 404


@app.route('/api/ocr_test', methods=['GET'])
def ocr_test():
    """OCR测试接口，无需上传文件，直接返回测试数据"""
    return jsonify({
        'success': True,
        'msg': 'OCR路由正常',
        'records': [{'name': '测试玩家', 'points': 12345}],
        'raw_count': 1,
        'lines': ['测试 12345']
    })

@app.route('/api/ocr_image', methods=['POST'])
def ocr_image():
    """OCR识别积分榜截图，返回结构化[{name, points}]"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '未上传文件'}), 400
    file = request.files['file']
    # 调试：写文件日志（绕过 nohup 缓冲）
    with open('/tmp/ocr_debug.log', 'a') as dbg:
        import time
        dbg.write(f"[OCR] {time.strftime('%H:%M:%S')} ocr_image() 被调用, file={file.filename}\n")
        dbg.flush()
    if file.filename == '':
        return jsonify({'success': False, 'error': '文件名为空'}), 400
    # 用时间戳+随机字符作文件名，避免 secure_filename 吞掉中文
    import time, uuid
    ext = os.path.splitext(file.filename)[1] or '.jpg'
    filename = f'ocr_{int(time.time())}_{uuid.uuid4().hex[:8]}{ext}'
    tmp_path = os.path.join(UPLOAD_DIR, filename)
    file.save(tmp_path)
    try:
        # 调用腾讯云OCR API
        print(f'[OCR Debug] 调用 tencent_ocr: {tmp_path}', flush=True)
        raw = tencent_ocr(tmp_path)
        print(f'[OCR Debug] tencent_ocr 返回 {len(raw) if raw else 0} 个结果', flush=True)
        if raw:
            for r in raw[:5]:
                print(f'[OCR Debug]   文字: {r[1]}, 置信度: {r[2]}')
        # 按行分组：Y中心相近的视为同一行（阈值=文字高度的一半）
        blocks = []
        for (bbox, text, conf) in raw:
            t = text.strip()
            if not t or conf < 30:
                continue
            y_center = (bbox[0][1] + bbox[2][1]) / 2
            x_center = (bbox[0][0] + bbox[2][0]) / 2
            blocks.append({'text': t, 'y': y_center, 'x': x_center})
        
        if not blocks:
            return jsonify({'success': True, 'records': [], 'text': ''})
        
        # 过滤垃圾文字块（UI按钮、标题等）
        junk_keywords = ['历史战绩', '历虫战绩', '我方排行', '敌方排行', '排名', '主公',
                         '个人积分', '军团', '查看参战', '转发截图', '战场介绍', '国派对',
                         '邀万', '查看', '参战名单', '活动', '奖励', '夺宝', '官渡', '邀约', '战场']
        blocks = [b for b in blocks if not any(kw in b['text'] for kw in junk_keywords)]
        
        if not blocks:
            return jsonify({'success': True, 'records': [], 'text': '', 'raw_count': 0, 'lines': []})
        
        # 策略：扫描所有文字块，找4+位数字（积分），
        # 对每个积分，找同行（±30px Y差距）内最右侧的非数字块作为名字
        y_tolerance = 30  # 同行判定阈值
        records = []
        used_names = set()  # 避免同一个名字块被重复使用
        
        # 找所有潜在积分块（4-12位纯数字）
        points_candidates = []
        for b in blocks:
            t = b['text'].strip()
            if re.match(r'^\d{4,12}$', t):
                points_candidates.append(b)
        
        # 按X坐标从右到左排序（优先匹配右侧积分）
        points_candidates.sort(key=lambda b: b['x'], reverse=True)
        
        for pts_b in points_candidates:
            pts = int(pts_b['text'].strip())
            # 找同行（Y差距<30px）内最右侧的非数字、非垃圾文字块
            same_row = [b for b in blocks
                        if abs(b['y'] - pts_b['y']) < y_tolerance
                        and b['x'] < pts_b['x']]
            if not same_row:
                continue
            same_row.sort(key=lambda b: b['x'], reverse=True)
            # 取最右侧的作为名字
            name_block = None
            for nb in same_row:
                t = nb['text'].strip()
                if re.match(r'^\d+$', t):  # 纯数字（排名）跳过
                    continue
                if any(kw in t for kw in junk_keywords):
                    continue
                name_block = nb
                break
            if not name_block:
                continue
            name = name_block['text'].strip()
            # 去掉排名前缀（如 "2 张三" → "张三"）
            name = re.sub(r'^\d{1,3}\s+', '', name).strip()
            if name and len(name) >= 2 and name not in used_names:
                used_names.add(name)
                records.append({'name': name, 'points': pts})
        
        full_text = '\n'.join(b['text'] for b in blocks)
        # 调试：返回每个文字块的详细信息（坐标、文本）
        raw_blocks_detail = [{'text': b['text'], 'y': round(b['y'], 1), 'x': round(b['x'], 1)} for b in blocks]
        return jsonify({'success': True, 'records': records, 'text': full_text, 'raw_count': len(blocks), 'raw_blocks': raw_blocks_detail})
    except Exception as e:
        print(f'[OCR] 识别失败: {e}')
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


# ============ 考勤API ============

# --- 成员管理 ---

@app.route('/api/attendance/members', methods=['GET'])
def list_members():
    """获取成员列表"""
    conn = sqlite3.connect(ATTENDANCE_DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    rows = c.execute('SELECT * FROM members ORDER BY name ASC').fetchall()
    conn.close()
    return jsonify({'success': True, 'data': [dict(r) for r in rows]})


@app.route('/api/attendance/members/import', methods=['POST'])
def import_members():
    """从属性表导入成员到考勤名单(前端传入名字列表)"""
    names = (request.json or {}).get('names', [])
    if not names:
        return jsonify({'success': False, 'error': '名字列表为空,请先在步骤1导入属性表'})
    conn = sqlite3.connect(ATTENDANCE_DB)
    c = conn.cursor()
    added = 0
    skipped = 0
    for name in names:
        name = name.strip()
        if not name:
            continue
        try:
            c.execute('INSERT INTO members (name) VALUES (?)', (name,))
            added += 1
        except sqlite3.IntegrityError:
            skipped += 1
    conn.commit()
    conn.close()
    log_attendance('IMPORT', f'导入成员: 新增{added}, 跳过{skipped}')
    return jsonify({'success': True, 'added': added, 'skipped': skipped})


@app.route('/api/attendance/members/add', methods=['POST'])
def add_member():
    """手动添加单个成员"""
    name = (request.json or {}).get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': '名字不能为空'})
    conn = sqlite3.connect(ATTENDANCE_DB)
    c = conn.cursor()
    try:
        c.execute('INSERT INTO members (name) VALUES (?)', (name,))
        conn.commit()
        log_attendance('ADD_MEMBER', name)
        ok = True
        err = None
    except sqlite3.IntegrityError:
        ok = False
        err = '成员已存在'
    conn.close()
    return jsonify({'success': ok, 'error': err})


@app.route('/api/attendance/members/delete', methods=['POST'])
def delete_member():
    """删除成员"""
    name = (request.json or {}).get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': '名字不能为空'})
    conn = sqlite3.connect(ATTENDANCE_DB)
    c = conn.cursor()
    c.execute('DELETE FROM members WHERE name = ?', (name,))
    conn.commit()
    conn.close()
    log_attendance('DEL_MEMBER', name)
    return jsonify({'success': True})


# --- 分配记录 ---

@app.route('/api/attendance/assignments', methods=['GET'])
def list_assignments():
    """获取所有分配记录"""
    conn = sqlite3.connect(ATTENDANCE_DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    rows = c.execute('SELECT * FROM assignments ORDER BY created_at DESC').fetchall()
    result = []
    for row in rows:
        d = dict(row)
        cnt = c.execute('SELECT role, COUNT(*) as cnt FROM assignment_members WHERE assignment_id=? GROUP BY role', (row['id'],)).fetchall()
        d['member_count'] = {r['role']: r['cnt'] for r in cnt}
        result.append(d)
    conn.close()
    return jsonify({'success': True, 'data': result})


@app.route('/api/attendance/assignments/save', methods=['POST'])
def save_assignment():
    """保存当前分配结果为一条分配记录"""
    data = request.json or {}
    section = data.get('section', '').strip()
    members_list = data.get('members', [])
    if not section:
        return jsonify({'success': False, 'error': '请选择军团'})
    conn = sqlite3.connect(ATTENDANCE_DB)
    c = conn.cursor()
    try:
        c.execute('INSERT OR IGNORE INTO assignments (section) VALUES (?)', (section,))
        conn.commit()
        aid = c.execute('SELECT id FROM assignments WHERE section=?', (section,)).fetchone()[0]
        c.execute('DELETE FROM assignment_members WHERE assignment_id=?', (aid,))
        for m in members_list:
            c.execute('INSERT INTO assignment_members (assignment_id, member_name, role, position) VALUES (?,?,?,?)',
                      (aid, m.get('name',''), m.get('role','正式'), m.get('position','')))
        conn.commit()
        log_attendance('SAVE_ASSIGN', f'{section} 共{len(members_list)}人')
        return jsonify({'success': True, 'assignment_id': aid})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)})
    finally:
        conn.close()


@app.route('/api/attendance/assignments/<int:aid>/members', methods=['GET'])
def get_assignment_members(aid):
    """获取某条分配记录的成员"""
    conn = sqlite3.connect(ATTENDANCE_DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    rows = c.execute('SELECT * FROM assignment_members WHERE assignment_id=? ORDER BY role, position', (aid,)).fetchall()
    conn.close()
    return jsonify({'success': True, 'data': [dict(r) for r in rows]})


# --- 战报管理 ---

@app.route('/api/attendance/reports', methods=['GET'])
def list_reports():
    """获取战报列表"""
    conn = sqlite3.connect(ATTENDANCE_DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    rows = c.execute('SELECT * FROM attendance_reports ORDER BY report_date DESC, section ASC').fetchall()
    result = []
    for row in rows:
        d = dict(row)
        stats = c.execute(
            'SELECT status, COUNT(*) as cnt FROM attendance_detail WHERE report_id=? GROUP BY status',
            (row['id'],)
        ).fetchall()
        d['status_count'] = {r['status']: r['cnt'] for r in stats}
        result.append(d)
    conn.close()
    return jsonify({'success': True, 'data': result})
@app.route('/api/attendance/reports/create', methods=['POST'])
def create_report():
    """创建战报:选择军团+日期+关联分配记录,OCR识别后自动标记出席并记录功勋积分"""
    data = request.json or {}
    section = data.get('section', '').strip()
    report_date = data.get('report_date', '').strip()
    assignment_id = data.get('assignment_id')
    ocr_data = data.get('ocr_data', [])

    # 兼容旧格式(只有名字列表)
    if isinstance(ocr_data, list) and len(ocr_data) > 0:
        if isinstance(ocr_data[0], str):
            ocr_data = [{'member_name': n, 'points': 0} for n in ocr_data]
        else:
            # 前端传的是 {name, points},转成后端格式
            ocr_data = [{'member_name': d.get('name', ''), 'points': d.get('points', 0)} for d in ocr_data]

    if not section or not report_date:
        return jsonify({'success': False, 'error': '请填写军团和日期'})

    conn = sqlite3.connect(ATTENDANCE_DB)
    c = conn.cursor()
    try:
        # 确保 points 列存在(向后兼容已有数据库)
        try:
            c.execute('ALTER TABLE attendance_detail ADD COLUMN points INTEGER DEFAULT 0')
            conn.commit()
        except Exception:
            pass

        c.execute('INSERT INTO attendance_reports (section, report_date) VALUES (?, ?)', (section, report_date))
        conn.commit()
        report_id = c.lastrowid

        if assignment_id:
            rows = c.execute('SELECT member_name, role FROM assignment_members WHERE assignment_id=?', (assignment_id,)).fetchall()
            base_members = {r[0]: r[1] for r in rows}
        else:
            rows = c.execute('SELECT name FROM members').fetchall()
            base_members = {r[0]: '正式' for r in rows}

        # 建立积分表(同名累加)
        pts_map = {}
        for d in ocr_data:
            name = d.get('member_name', d.get('name', ''))
            pts = int(d.get('points', 0) or 0)
            if name:
                pts_map[name] = pts

        ocr_set = set(pts_map.keys())

        for name, role in base_members.items():
            if name in ocr_set:
                status = '候补上场' if role == '候补' else '出席'
            else:
                status = '缺席'
            pts = pts_map.get(name, 0)
            c.execute('INSERT INTO attendance_detail (report_id, member_name, status, points) VALUES (?,?,?,?)',
                      (report_id, name, status, pts))

        for name, pts in pts_map.items():
            if name not in base_members:
                c.execute('INSERT INTO attendance_detail (report_id, member_name, status, points) VALUES (?,?,?,?)',
                          (report_id, name, '候补上场', pts))

        conn.commit()
        log_attendance('CREATE_REPORT', f'{section} {report_date} ID:{report_id} 积分记录:{len(pts_map)}人')
        return jsonify({'success': True, 'report_id': report_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        conn.close()


@app.route('/api/attendance/reports/delete', methods=['POST'])
def delete_report():
    """删除战报"""
    report_id = (request.json or {}).get('report_id')
    if not report_id:
        return jsonify({'success': False, 'error': '缺少report_id'})
    conn = sqlite3.connect(ATTENDANCE_DB)
    c = conn.cursor()
    c.execute('DELETE FROM attendance_detail WHERE report_id=?', (report_id,))
    c.execute('DELETE FROM attendance_reports WHERE id=?', (report_id,))
    conn.commit()
    conn.close()
    log_attendance('DEL_REPORT', f'ID:{report_id}')
    return jsonify({'success': True})


# --- 考勤明细 ---

@app.route('/api/attendance/detail/<int:report_id>', methods=['GET'])
def get_attendance_detail(report_id):
    """获取某战报的考勤明细"""
    conn = sqlite3.connect(ATTENDANCE_DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    rows = c.execute('SELECT * FROM attendance_detail WHERE report_id=? ORDER BY status, member_name', (report_id,)).fetchall()
    report = c.execute('SELECT * FROM attendance_reports WHERE id=?', (report_id,)).fetchone()
    conn.close()
    return jsonify({
        'success': True,
        'report': dict(report) if report else None,
        'data': [dict(r) for r in rows]
    })


@app.route('/api/attendance/detail/update', methods=['POST'])
def update_attendance_status():
    """更新成员考勤状态"""
    data = request.json or {}
    report_id = data.get('report_id')
    member_name = data.get('member_name', '').strip()
    status = data.get('status', '').strip()
    if not all([report_id, member_name, status]):
        return jsonify({'success': False, 'error': '参数不完整'})
    if status not in ('出席', '缺席', '请假', '候补上场'):
        return jsonify({'success': False, 'error': f'无效状态: {status}'})
    conn = sqlite3.connect(ATTENDANCE_DB)
    c = conn.cursor()
    c.execute('UPDATE attendance_detail SET status=?, updated_at=CURRENT_TIMESTAMP WHERE report_id=? AND member_name=?',
              (status, report_id, member_name))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/attendance/detail/batch_update', methods=['POST'])
def batch_update_status():
    """批量更新考勤状态"""
    data = request.json or {}
    report_id = data.get('report_id')
    updates = data.get('updates', [])
    if not report_id or not updates:
        return jsonify({'success': False, 'error': '参数不完整'})
    conn = sqlite3.connect(ATTENDANCE_DB)
    c = conn.cursor()
    for u in updates:
        status = u.get('status', '')
        if status not in ('出席', '缺席', '请假', '候补上场'):
            continue
        c.execute('UPDATE attendance_detail SET status=?, updated_at=CURRENT_TIMESTAMP WHERE report_id=? AND member_name=?',
                  (status, report_id, u.get('member_name', '')))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'updated': len(updates)})


# --- 导出 ---

@app.route('/api/attendance/export', methods=['POST'])
def export_attendance():
    """导出战报Excel"""
    data = request.json or {}
    password = data.get('password', '')
    report_id = data.get('report_id')
    if password != 'lingxiao2026':
        return jsonify({'success': False, 'error': '密码错误'}), 403
    if not report_id:
        return jsonify({'success': False, 'error': '请选择战报'})

    conn = sqlite3.connect(ATTENDANCE_DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    report = c.execute('SELECT * FROM attendance_reports WHERE id=?', (report_id,)).fetchone()
    if not report:
        conn.close()
        return jsonify({'success': False, 'error': '战报不存在'})
    rows = c.execute('SELECT * FROM attendance_detail WHERE report_id=?', (report_id,)).fetchall()
    conn.close()

    from io import BytesIO
    wb = openpyxl.Workbook()

    status_order = {'出席': 0, '请假': 1, '候补上场': 2, '缺席': 3}
    sorted_rows = sorted(rows, key=lambda r: (status_order.get(r['status'], 9), r['member_name']))

    ws = wb.active
    ws.title = f'{report["section"]} 考勤'
    ws.append([f'{report["section"]} 考勤战报 - {report["report_date"]}'])
    ws.merge_cells('A1:C1')
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(['姓名', '状态', '功勋积分', '更新时间'])
    for r in sorted_rows:
        ws.append([r['member_name'], r['status'], r['points'] or 0, r['updated_at'] or ''])
    ws.append([])
    ws.append(['统计'])
    total_pts = sum(r['points'] or 0 for r in rows)
    for st in ['出席', '请假', '候补上场', '缺席']:
        cnt = sum(1 for r in rows if r['status'] == st)
        ws.append([st, cnt])
    ws.append(['功勋积分合计', total_pts])

    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 18

    absent = [r for r in sorted_rows if r['status'] == '缺席']
    if absent:
        ws2 = wb.create_sheet('缺席名单')
        ws2.append(['缺席成员'])
        ws2['A1'].font = Font(bold=True)
        for r in absent:
            ws2.append([r['member_name']])
        ws2.column_dimensions['A'].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'考勤战报_{report["section"]}_{report["report_date"]}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=filename)


# ============ 主程序 ============

import threading

def preload_ocr_background():
    """后台线程预加载OCR模型"""
    import time
    time.sleep(2)  # 等Flask启动后再加载
    print('[OCR] 后台预加载OCR模型...')
    try:
        get_ocr_reader()
        print('[OCR] OCR模型加载成功,可以使用')
    except Exception as e:
        print(f'[OCR] OCR模型加载失败: {e}')

# ============ 队形配置模块 ============

# 预设经典队形
DEFAULT_FORMATIONS = [
    {
        'id': 'classic_6team',
        'name': '经典6队(9人)',
        'teamCount': 6,
        'teamSize': 9,
        'hasAB': True,
        'groupASize': 4,
        'tasks': {
            '0-10': ['占工匠坊', '占兵器坊', '占自己方上粮仓', '占自己方中粮仓', '占自己方下粮仓', '占乌巢'],
            '10-20': ['占官渡', '占霹雳车', '占大粮仓', '驻防采集', '占乌巢', '占官渡'],
            '20+': ['驻防采集', '驻防采集', '驻防采集', '驻防采集', '驻防采集', '驻防采集'],
            '0-10-b': ['占工匠坊', '占兵器坊', '占自己方上粮仓', '占自己方中粮仓', '占自己方下粮仓', '占乌巢'],
            '10-20-b': ['占官渡', '占霹雳车', '占大粮仓', '驻防采集', '占乌巢', '占官渡'],
            '20+-b': ['驻防采集', '驻防采集', '驻防采集', '驻防采集', '驻防采集', '驻防采集']
        }
    },
    {
        'id': 'quick_5team',
        'name': '快速5队(6人)',
        'teamCount': 5,
        'teamSize': 6,
        'hasAB': True,
        'groupASize': 3,
        'tasks': {
            '0-10': ['占工匠坊', '占兵器坊', '占自己方上粮仓', '占乌巢', '占官渡'],
            '10-20': ['占官渡', '占霹雳车', '占大粮仓', '驻防采集', '占乌巢'],
            '20+': ['驻防采集', '驻防采集', '驻防采集', '驻防采集', '驻防采集'],
            '0-10-b': ['占工匠坊', '占兵器坊', '占自己方上粮仓', '占乌巢', '占官渡'],
            '10-20-b': ['占官渡', '占霹雳车', '占大粮仓', '驻防采集', '占乌巢'],
            '20+-b': ['驻防采集', '驻防采集', '驻防采集', '驻防采集', '驻防采集']
        }
    },
    {
        'id': 'heavy_8team',
        'name': '重兵8队(4人)',
        'teamCount': 8,
        'teamSize': 4,
        'hasAB': True,
        'groupASize': 2,
        'tasks': {
            '0-10': ['占工匠坊', '占兵器坊', '占自己方上粮仓', '占自己方中粮仓', '占自己方下粮仓', '占乌巢', '占官渡', '占霹雳车'],
            '10-20': ['占官渡', '占霹雳车', '占大粮仓', '驻防采集', '占乌巢', '占官渡', '驻防采集', '驻防采集'],
            '20+': ['驻防采集'] * 8,
            '0-10-b': ['占工匠坊', '占兵器坊', '占自己方上粮仓', '占自己方中粮仓', '占自己方下粮仓', '占乌巢', '占官渡', '占霹雳车'],
            '10-20-b': ['占官渡', '占霹雳车', '占大粮仓', '驻防采集', '占乌巢', '占官渡', '驻防采集', '驻防采集'],
            '20+-b': ['驻防采集'] * 8
        }
    }
]

# 预设任务选项
TASK_OPTIONS = [
    '占工匠坊', '占兵器坊',
    '占自己方上粮仓', '占自己方中粮仓', '占自己方下粮仓',
    '占乌巢', '占官渡', '占霹雳车', '占大粮仓',
    '驻防采集',
    '防守反击', '支援邻队', '待命'
]

@app.route('/api/formations/presets')
def get_formation_presets():
    """获取预设队形列表"""
    return jsonify({'success': True, 'formations': DEFAULT_FORMATIONS, 'tasks': TASK_OPTIONS})

@app.route('/api/formations/validate', methods=['POST'])
def validate_formation():
    """验证队形配置是否有效"""
    data = request.get_json()
    team_count = data.get('teamCount', 6)
    team_size = data.get('teamSize', 9)
    has_ab = data.get('hasAB', True)
    group_a_size = data.get('groupASize', 0)
    
    total = team_count * team_size
    if has_ab:
        # A/B组模式：队长1人 + A组 + B组
        if group_a_size < 1 or group_a_size > team_size - 2:
            return jsonify({'valid': False, 'error': f'A组人数需在 1~{team_size-2} 之间'})
        group_b_size = team_size - 1 - group_a_size
        return jsonify({
            'valid': True,
            'totalMembers': total,
            'teamCount': team_count,
            'teamSize': team_size,
            'hasAB': has_ab,
            'captainCount': team_count,
            'groupASize': group_a_size,
            'groupBSize': group_b_size
        })
    
    return jsonify({
        'valid': True,
        'totalMembers': total,
        'teamCount': team_count,
        'teamSize': team_size,
        'hasAB': has_ab
    })

# ============ 版本管理路由 ============
import subprocess, os as _os

_VERSION_REPO = os.path.dirname(os.path.abspath(__file__))

@app.route('/version')
def version_page():
    """版本管理页面"""
    return render_template('version.html')

@app.route('/api/version/log')
def version_log():
    """获取 Git 提交历史"""
    try:
        result = subprocess.run(
            ['git', 'log', '--pretty=format:%H|%an|%ae|%ad|%s', '--date=short', '--all'],
            cwd=_VERSION_REPO, capture_output=True, text=True, timeout=10
        )
        if result.returncode != 0:
            return jsonify({'success': False, 'error': result.stderr})
        commits = []
        for line in result.stdout.strip().split('\n'):
            if not line:
                continue
            parts = line.split('|', 4)
            if len(parts) == 5:
                commits.append({
                    'hash': parts[0][:7],
                    'full_hash': parts[0],
                    'author': parts[1],
                    'email': parts[2],
                    'date': parts[3],
                    'message': parts[4]
                })
        head = subprocess.run(['git', 'rev-parse', '--short', 'HEAD'],
                             cwd=_VERSION_REPO, capture_output=True, text=True)
        return jsonify({'success': True, 'commits': commits, 'head': head.stdout.strip()})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/version/diff/<commit_hash>')
def version_diff(commit_hash):
    """查看某次提交的 diff"""
    try:
        result = subprocess.run(
            ['git', 'show', '--stat', commit_hash],
            cwd=_VERSION_REPO, capture_output=True, text=True, timeout=10
        )
        return jsonify({'success': True, 'diff': result.stdout})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/version/rollback', methods=['POST'])
def version_rollback():
    """回滚到指定 commit（需密码确认）"""
    data = request.get_json()
    commit_hash = data.get('commit_hash', '').strip()
    password = data.get('password', '').strip()
    if password != '334dengni':
        return jsonify({'success': False, 'error': '密码错误'})
    if not commit_hash:
        return jsonify({'success': False, 'error': '缺少 commit_hash'})
    try:
        subprocess.run(['git', 'add', '-A'], cwd=_VERSION_REPO, capture_output=True)
        subprocess.run(['git', 'stash'], cwd=_VERSION_REPO, capture_output=True)
        result = subprocess.run(
            ['git', 'checkout', commit_hash],
            cwd=_VERSION_REPO, capture_output=True, text=True, timeout=10
        )
        if result.returncode != 0:
            subprocess.run(['git', 'stash', 'pop'], cwd=_VERSION_REPO, capture_output=True)
            return jsonify({'success': False, 'error': result.stderr})
        return jsonify({'success': True, 'message': f'已回滚到 {commit_hash[:7]}'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/version/restore_head', methods=['POST'])
def version_restore_head():
    """恢复到最新版本"""
    try:
        result = subprocess.run(
            ['git', 'checkout', 'main'],
            cwd=_VERSION_REPO, capture_output=True, text=True, timeout=10
        )
        return jsonify({'success': True, 'message': '已恢复到最新版本'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


if __name__ == '__main__':
    # 启动后台线程预加载OCR(不阻塞主服务)
    t = threading.Thread(target=preload_ocr_background, daemon=True)
    t.start()
    print('[启动] Flask服务启动中... (OCR模型将在后台加载)')
    app.run(debug=False, host='0.0.0.0', port=5001, threaded=True)
